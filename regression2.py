import sys
import numpy as np
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
from collections import OrderedDict
import dateutil.relativedelta
from openpyxl import load_workbook

class chip:
    def __init__(self, name, teir, TOdate, scalar):
        self.name = name
        self.tier = teir
        self.TO = TOdate
        self.scalar = scalar
       

def main():
    workbook = load_workbook("ForecastModel.xlsx")
    profile_sheet = workbook["Profiles"]
    #create profile nested dictionary {tier:{days from TO:usage}}
    profile_dict = {}
    for row in profile_sheet.iter_rows():
        to_dict = {}
        start = -9
        for col in range(1,16):
            if row[col].value == "TO-9":
                break
            elif row[col].value == None:
                to_dict[start] = 0
            else:
                to_dict[start] = row[col].value
            start +=1
            profile_dict[row[0].value] = to_dict

     #create array of chips to be made       
    roadmap_sheet = workbook["Chip Roadmap"]
    chips = []
    for row in roadmap_sheet.iter_rows():
        if row[0].value == "Chip":
            continue
        else:
            name = row[0].value
            tier = row[1].value
            date = row[2].value
            scaler = row[3].value
            temp = chip(name, tier, date, scaler)
            chips.append(temp)
    
    #x axis min(date from chips[])-9 - max(date from chips[])+5 !hard code alert!
    maxx = max(chip.TO for chip in chips) + dateutil.relativedelta.relativedelta(months=5)
    minx = min(chip.TO for chip in chips) - dateutil.relativedelta.relativedelta(months=9)
    #fill xaxis list
    start = minx
    end = maxx
    total_months = lambda dt: dt.month + 12 * dt.year
    xaxis = []
    for tot_m in range(total_months(start)-1, total_months(end)):
        y, m = divmod(tot_m, 12)
        xaxis.append(datetime(y, m+1, 1))
    

    
    #y axis nessted array
    yax = []
    #loop through chips to be made
    for i in chips:
        temp = []
        #loop through x axis
        for j in xaxis:
            
            if j == i.TO - dateutil.relativedelta.relativedelta(months=9): #hard code alert
                for k in range(-9,6):
                    temp.append(profile_dict[i.tier][k]*i.scalar)
                
            elif j< i.TO - dateutil.relativedelta.relativedelta(months=9) or j> i.TO + dateutil.relativedelta.relativedelta(months=5): #hard code alert
                temp.append(0)
        
        yax.append(temp)
    
    
    #plot!
    #get list of names of chips to be made
    names = []
    for x in chips:
        names.append(x.name)
  
    plt.stackplot(xaxis,yax, labels = names)
    plt.legend(loc='upper left')
    plt.show()

        
   

   

if __name__ == "__main__":
    main()