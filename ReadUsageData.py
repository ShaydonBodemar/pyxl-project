# Module Name: ReadUsageData.py
# Author: Shaydon Bodemar
# Date: 21 September 2020
# Overview: This module will be primarily responsible for inputting the data from the excel sheet for all model parameters.

from openpyxl import load_workbook
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
import numpy as np
from numpy import ma
import matplotlib
import matplotlib.pyplot as plt

class ReadUsageData:
    """
    @brief Constructor for this module
    @param excel_filename Used to determine the file that will be interacted with.
    @note Only opens the file to be used and initializes the member variables.
    """
    def __init__(self, excel_filename):
        self._workbook = load_workbook(excel_filename)  # The workbook opened using openpyxl library (cannot be read_only=True to allow cell.col_idx)
        self._metadata_fields = []                      # Array of fields for chip metadata
        self._chip_profiles = {}                        # Nested Dictionary to map each chip to its metadata by name
        self._tapeout_offset = 0                        # Offset of the TO (tapeout) date from intial value (time prior to TO)
        self._tool_usage_by_tier = {}                   # Dictionary of arrays of tool usage data for each tier of chip
    

    """
    @brief Destructor for this module
    @note Closes workbook used for reading sheets in this module.
    """
    def __del__(self):
        self._workbook.close()


    """
    @brief Reads the metadata for each chip (scaler, TO date, etc)
    @param worksheet_name Name of the worksheet to be used for reading in appropriate data
    @note Wipes any metadata currently stored
    TODO: Handle incorrect/nonexistent worksheet_name
    """
    def ReadChipMetaData(self, worksheet_name):
        cur_worksheet = self._workbook[worksheet_name]

        # Reads in the data labels for the first row to be used later as keys in dict
        self._metadata_fields = []
        for cell in cur_worksheet[1]:
            self._metadata_fields.append(cell.value)

        # Reads in all data fields to create a full nested dictionary representing each chip's metadata
        self._chip_profiles = {}
        for row in cur_worksheet.iter_rows(min_row=2):
            chip_name = row[0].value
            self._chip_profiles[chip_name] = {}
            for cell in range(1,len(row)):
                self._chip_profiles[chip_name][self._metadata_fields[cell]] = row[cell].value


    """
    @brief Reads the tool usage profiles for each tier
    @param worksheet_name Name of the worksheet to be used for reading in appropriate data
    @note Overwites any tier data previously present
    TODO: Handle incorrect/nonexistent worksheet_name
    NOTE: TO offset conditional could be handled dynamically if parameter is taken
    """
    def ReadChipTierData(self, worksheet_name):
        cur_worksheet = self._workbook[worksheet_name]

        # Determines the offset of the TO date
        for cell in cur_worksheet[1]:
            if cell.value == 'TO':
                self._tapeout_offset = cell.col_idx - 2     # subtract two to account for non-data first column and 0 starting index

        # Reads in all tool usage data and maps it to the appropriate tier
        self._tool_usage_by_tier = {}
        for row in cur_worksheet.iter_rows(min_row=2):
            tier_name = row[0].value
            self._tool_usage_by_tier[tier_name] = []
            for cell in range(1,len(row)):
                if row[cell].value == None:
                    self._tool_usage_by_tier[tier_name].append(0)
                else:
                    self._tool_usage_by_tier[tier_name].append(row[cell].value)

    
    """
    @brief Writes all scaled profile usage data for each chip
    @param worksheet_name Name of the worksheet to be used for writing data
    NOTE: Handle worksheet_name that already exists to prevent overwriting
    """
    def WriteAllProfileData(self, worksheet_name):
        return 1


    """
    @brief Finds the earliest date on which there will be tool usage based on the data
    TODO: Handle edge cases (unpopulated data structures)
    TODO: Handle dynamic searching for 'TO Date' field
    TODO: Account for empty begin dates
    """
    def EarliestDateOfConcern(self):
        earliest = datetime.max
        months_prior = relativedelta(months=self._tapeout_offset)
        for chip in self._chip_profiles:
            if (self._chip_profiles[chip]['TO Date'] - months_prior) < earliest:
                earliest = self._chip_profiles[chip]['TO Date'] - months_prior
        return earliest

    
    """
    @brief Finds the latest date on which there will be tool usage based on the data
    TODO: Handle edge cases (unpopulated data structures)
    TODO: Handle dynamic searching for 'TO Date' field
    TODO: Account for empty end dates
    """
    def LatestDateOfConcern(self):
        latest = datetime.min
        months_offset = len(self._tool_usage_by_tier[next(iter(self._tool_usage_by_tier))]) - self._tapeout_offset
        months_prior = relativedelta(months=months_offset)
        for chip in self._chip_profiles:
            if (self._chip_profiles[chip]['TO Date'] + months_prior) > latest:
                latest = self._chip_profiles[chip]['TO Date'] + months_prior
        return latest

    
    """
    @brief Creates an array of all months between a given start and end date
    @param begin_date Represents the date the array will start at
    @param end_date Represents the date the array will end at
    """
    def DateRangeList(self, begin_date, end_date):
        cur_date = begin_date
        date_range = []
        while cur_date <= end_date:
            date_range.append(cur_date)
            cur_date += relativedelta(months=1)
        return date_range


    """
    @brief Creates the plot of the data once read in
    TODO: Ensure data is actually present for required fields
    TODO: Dynamic handling of field names
    NOTE: Typo for Scaler
    """
    def PlotData(self):
        chip_names = []
        for chip in self._chip_profiles:
            chip_names.append(chip)

        date_range = self.DateRangeList(self.EarliestDateOfConcern(), self.LatestDateOfConcern())

        # TODO: Fill to size of full arrays
        all_data = []
        for chip in chip_names:
            full_offset = date_range.index(self._chip_profiles[chip]['TO Date']) - self._tapeout_offset
            print(full_offset)
            profile = np.array(self._tool_usage_by_tier[self._chip_profiles[chip]['Tier']])
            profile = np.multiply(profile, self._chip_profiles[chip]['Scaler'])
            profile = np.pad(profile, (full_offset, len(date_range)-full_offset-profile.size), 'constant')
            all_data.append(profile.tolist())
        # TODO: remove after debugging
        print(all_data)
        # TODO: remove after debugging
        print(len(all_data[0]) == len(date_range))

        plt.stackplot(date_range, all_data, labels=chip_names)
        plt.legend(loc='upper left')
        plt.savefig('stacked_area_chart.png')